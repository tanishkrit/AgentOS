from openpyxl import load_workbook
from pathlib import Path

workspace = Path(r"c:\Users\kritg\OneDrive\Desktop\Tanish\Promgrams\AI Agents\workspace")
# Find the latest agent_results file
files = sorted(list(workspace.glob("agent_results_*.xlsx")))

if not files:
    print("No Excel files found.")
else:
    latest = files[-1]
    print(f"Reading Excel file: {latest.name}")
    
    wb = load_workbook(latest)
    ws = wb.active
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < 15: # Print first 15 rows
            print(f"Row {row_idx+1}: {row}")
        else:
            break
