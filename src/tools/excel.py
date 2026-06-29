"""
Excel Tool — Direct Excel File Creation via openpyxl

Creates .xlsx files directly without needing to control Excel via keyboard.
Can also open the created file in Excel for the user to see.

This is more reliable than keyboard-driven Excel automation for data entry.
"""

import logging
import os
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.config import Config

logger = logging.getLogger(__name__)


class ExcelTool:
    """
    Creates professional Excel workbooks directly using openpyxl.
    """

    def create_workbook(
        self,
        headers: list[str],
        rows: list[list[str]],
        filename: str | None = None,
        sheet_name: str = "Results",
    ) -> str:
        """
        Create a styled Excel workbook with headers and data rows.

        Args:
            headers: Column header names.
            rows: List of row data (each row is a list of cell values).
            filename: Output filename (auto-generated if None).
            sheet_name: Name for the worksheet.

        Returns:
            Absolute path to the created .xlsx file.
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"agent_results_{timestamp}.xlsx"

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        # Save to workspace directory
        output_path = Config.WORKSPACE_ROOT / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # ── Styles ───────────────────────────────────────────────────
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Calibri", size=11)
        data_alignment = Alignment(vertical="top", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # ── Write headers ────────────────────────────────────────────
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Write data rows ──────────────────────────────────────────
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                # Alternate row coloring
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # ── Auto-fit column widths ───────────────────────────────────
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 60))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 4

        # ── Freeze header row ────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Save ─────────────────────────────────────────────────────
        wb.save(str(output_path))
        logger.info(f"Excel workbook saved: {output_path}")
        return str(output_path)

    def open_in_excel(self, filepath: str) -> None:
        """
        Open an Excel file in the default spreadsheet application.

        Uses os.startfile on Windows for native file association.
        """
        path = Path(filepath)
        if not path.exists():
            logger.error(f"File not found: {filepath}")
            return

        try:
            logger.info(f"Opening file in Excel: {filepath}")
            os.startfile(str(path))
        except Exception as e:
            logger.error(f"Failed to open file: {e}")
            # Fallback: try subprocess
            try:
                subprocess.Popen(["start", "", str(path)], shell=True)
            except Exception as e2:
                logger.error(f"Fallback open also failed: {e2}")
